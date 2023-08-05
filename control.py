import openpyxl
from openpyxl import Workbook

def write_to_excel(data):
    try:
        # Mở file Excel nếu đã tồn tại, nếu không tạo mới
        workbook = openpyxl.load_workbook('data.xlsx')
    except FileNotFoundError:
        # Tạo một file Excel mới nếu chưa tồn tại
        workbook = Workbook()
        workbook.save('data.xlsx')

    # Chọn sheet đầu tiên
    sheet = workbook.active

    # Tìm hàng tiếp theo trống
    next_row = sheet.max_row + 1

    for i, value in enumerate(['STT','Dang Vien','Ten Thuoc','Ngay Nhap','Hoat Chat','Don vi san xuat','Dia chi'], start=1):
        sheet.cell(row = 1, column=i).value = value

    # Ghi dữ liệu vào hàng tiếp theo
    for i, value in enumerate(data, start=1):
        if i==1:
            sheet.cell(row=next_row, column=i).value = next_row - 1
        else : 
            sheet.cell(row=next_row, column=i).value = value
        

    # Lưu file Excel
    workbook.save('data.xlsx')

def write_to_excel_out_of_date(data):
    try:
        # Mở file Excel nếu đã tồn tại, nếu không tạo mới
        workbook = openpyxl.load_workbook('out_of_date.xlsx')
    except FileNotFoundError:
        # Tạo một file Excel mới nếu chưa tồn tại
        workbook = Workbook()
        workbook.save('out_of_date.xlsx')

    # Chọn sheet đầu tiên
    sheet = workbook.active

    # Tìm hàng tiếp theo trống
    next_row = sheet.max_row + 1

    for i, value in enumerate(['STT','Dang Vien','Ten Thuoc','Ngay Nhap','Hoat Chat','Don vi san xuat','Dia chi'], start=1):
        sheet.cell(row = 1, column=i).value = value

    # Ghi dữ liệu vào hàng tiếp theo
    for i, value in enumerate(data, start=1):
        sheet.cell(row=next_row, column=i).value = value

    # Lưu file Excel
    workbook.save('out_of_date.xlsx')

